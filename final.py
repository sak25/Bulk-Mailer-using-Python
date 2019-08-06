import smtplib, ssl,csv

from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from email import encoders
import getpass
import logging
from openpyxl import load_workbook
import re


def num(x):       # function to convert alphabet to numerical value
    lower=x.lower()
    no=ord(lower)
    result=no-96
    return result
    
# set file path
filepath="users.xlsx"        #Excel file should be in the same directory as script
# load demo.xlsx 
wb=load_workbook(filepath)
# select sheet which is active
sheet=wb.active
print("Welcome to Bulk Mail Software.")
print("")
print("General instructions to use:")
print("")
print("1.To Start sending bulk mails make an excel file with name users.xlsx with column names like To,Bcc,Cc,Subject,Attachment etc in its first row and save it in the same directory.")
print("2.Make sure that the cell containg attachment has file's full address with extension like Ex. C:/Users/Desktop/document.pdf")
print("3.Make sure you allow access to less secure apps in your email address settings to send mails.")
print("4.Make a new file with the name logs.log to store log status of mails in the same directory.")
print("5.Make a text file with name body.txt to store body of mail in the same directory.")
print("6.For customisation in body of message enclose column name of excel file in the format like @B@ mandatorily where B is column no. of field")
print("7.To add server address and port no for sending email from simply enter them in a file with name configurations.csv")
# get max column count
to_column=0
    
bcc_column=0

cc_column=0

subject_column=0

attachment_column=0

max_column=sheet.max_column

for j in range(1,max_column+1):
    if str(sheet.cell(row=1,column=j).value.lower())=="to":
        to_column+=j
    if str(sheet.cell(row=1,column=j).value.lower())== "bcc":
        bcc_column+=j
    if str(sheet.cell(row=1,column=j).value.lower())=="cc":
        cc_column+=j
    if str(sheet.cell(row=1,column=j).value.lower())=="subject":
        subject_column+=j
    if str(sheet.cell(row=1,column=j).value.lower())== "attachment":
        attachment_column+=j
    else:
        continue
print("")
null = input("To start using press enter")
print("")
sender_email=input("Enter Sender's Email Address:")
password =input("Enter Password:") 
print("")    # for Password in encrypted form use: password=getpass.getpass()

print("Please Wait,it may take few seconds.......")
print("")

logging.basicConfig(filename='logs.log', filemode='a',level=logging.DEBUG,format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')


context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:

    try:



        server.login(sender_email, password)

        logging.info('Server connection Successful')  # Skip header row


        max_row=sheet.max_row
        
        for l in range(2,max_row+1):    
            fp=open('body.txt',"r+")
                    
            content=fp.read()
            f=open('body.txt',"r")
            line=f.readline()
            
            
            
            while line:
                
                for word in line.split():
                
                    if (re.match(r'@[A-Za-z]@',str(word)) is not None ):
                        result=re.match(r'@[A-Za-z]@',word)
                        store=result.group(0)
                        
                    
                        cleanString = re.sub(r'@','', store)
                        value=num(cleanString)
                        value=int(value)
                        new=(sheet.cell(row=l,column=value).value)
                        
                        content=content.replace(word,new)
                        fp.close()
                    else:
                        continue
                line=f.readline()    
            
            f.close()
            fp.close() 
            
            body=content
            To=str(sheet.cell(row=l,column=to_column).value)
            Bcc=str(sheet.cell(row=l,column=bcc_column).value)
            Cc=str(sheet.cell(row=l,column=cc_column).value)
            
           



            
        # Creating a multipart message and setting headers
            message = MIMEMultipart()
            message["From"] = sender_email
            message["To"] = To
            message["Subject"] = sheet.cell(row=l,column=subject_column).value
            message["Bcc"] = Bcc
            message["Cc"] = Cc
            message.attach(MIMEText(body, "plain"))# for Mass mailing with different specifications

            if((sheet.cell(row=l,column=attachment_column).value)!= None):           
                
            
                filename = str(sheet.cell(row=l,column=attachment_column).value)
                # attachment file In same directory as  python script
                        # Opening PDF file in binary mode
                with open(filename, "rb") as attachment:
               

                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

            # Encoding file in ASCII characters to send by email
                    encoders.encode_base64(part)

                # Adding header as key/value pair to attachment part
                    part.add_header("Content-Disposition", f"attachment; filename= {filename}")
                

                # Adding attachment to message and converting message to string
                    message.attach(part)
                
            # Logging in to server using secure context and sending email
            
            
       
            try:

                    
                server.sendmail(message["From"], [To, Bcc, Cc], message.as_string())
        
        
        

                logging.info('Recipient:%s: Status:Success',To)
                            
        
            except:

        
                logging.warning('Recipient:%s: Status:Failed',To)
                                                    

        print("Thank You For Using Bulk Mailer!! ")

        print("Kindly,check logs for mail status") 
        print("")

    except:

        logging.critical('Server connection Unsuccessful: Try Again') 
        print("Enter correct ID and password or allow less secure apps the access, Please try Again!!")
        
   
       
 
    
        
        